# requirements: flask, flask-cors, openpyxl

# app.py
from flask import Flask, request, send_file, jsonify, render_template
from flask_cors import CORS
from openpyxl import Workbook, load_workbook  # Certifique-se de importar o método correto
import io
from threading import Lock
import os
import json

app = Flask(__name__)
CORS(app)

# Store data in memory for simplicity
data_rows = []
data_lock = Lock()

# Define the expected columns for the spreadsheet
columns = [
    "Nº da Torre",
    "Nº da Ordem",
    "Especialidade",
    "LT",
    "Parte Afetada (Macro)",
    "Parte Afetada (Micro)",
    "Sintomas Dano (Macro)",
    "Sintomas Dano (Micro)",
    "Data Inspeção",
    "Data Nec. Exec.",
    "Prioridade",
    "Texto do Item",
    "Observação (Texto da Situação)"
]

# Variáveis globais para armazenar os valores padrão
default_values = {
    "Nº da Ordem": None,
    "Especialidade": None,
    "LT": None
}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    global default_values
    data = request.json

    # Reutiliza os valores padrão se os campos não forem enviados
    if "Nº da Ordem" in data and data["Nº da Ordem"]:
        default_values["Nº da Ordem"] = data["Nº da Ordem"]
    else:
        data["Nº da Ordem"] = default_values["Nº da Ordem"]

    if "Especialidade" in data and data["Especialidade"]:
        default_values["Especialidade"] = data["Especialidade"]
    else:
        data["Especialidade"] = default_values["Especialidade"]

    if "LT" in data and data["LT"]:
        default_values["LT"] = data["LT"]
    else:
        data["LT"] = default_values["LT"]

    # Verifica se todos os campos obrigatórios estão presentes
    required_fields = [key for key in columns if key not in ["Nº da Ordem", "Especialidade", "LT"]]
    if all(key in data for key in required_fields):
        data_rows.append(data)
        save_data_to_file()  # Salva os dados no arquivo JSON
        return jsonify({"status": "success"})
    else:
        return jsonify({"status": "error", "message": "Invalid data format"}), 400

def generate_excel():
    # Caminho para o modelo de planilha
    modelo_path = "modelo.xlsx"  # Certifique-se de que o arquivo modelo.xlsx está no mesmo diretório do script

    # Verifica se o modelo existe
    if not os.path.exists(modelo_path):
        raise FileNotFoundError("O arquivo modelo.xlsx não foi encontrado.")

    # Carrega o modelo de planilha existente
    try:
        wb = load_workbook(modelo_path)
        ws = wb.active
    except Exception as e:
        raise Exception(f"Erro ao carregar o modelo de planilha: {str(e)}")

    # Preenche os textos de "Nº da Ordem", "Especialidade" e "LT" no cabeçalho
    ws["C3"] = default_values["Nº da Ordem"] or ""  # Preenche o valor de "Nº da Ordem"
    ws["J3"] = default_values["Especialidade"] or ""  # Preenche o valor de "Especialidade"
    ws["Q3"] = default_values["LT"] or ""  # Preenche o valor de "LT"

    # Preenche os dados a partir da linha desejada
    start_row = 6  # Supondo que os dados começam na linha 6
    for i, row in enumerate(data_rows, start=start_row):
        try:
            ws.cell(row=i, column=1, value=row.get("Nº da Torre", ""))  # Coluna A
            ws.cell(row=i, column=2, value=row.get("Parte Afetada (Macro)", ""))  # Coluna B
            ws.cell(row=i, column=3, value=row.get("Parte Afetada (Micro)", ""))  # Coluna C
            ws.cell(row=i, column=4, value=row.get("Sintomas Dano (Macro)", ""))  # Coluna D
            ws.cell(row=i, column=5, value=row.get("Sintomas Dano (Micro)", ""))  # Coluna E
            ws.cell(row=i, column=6, value=row.get("Data Inspeção", ""))  # Coluna F
            ws.cell(row=i, column=8, value=row.get("Data Nec. Exec.", ""))  # Coluna G
            ws.cell(row=i, column=10, value=row.get("Prioridade", ""))  # Coluna H
            ws.cell(row=i, column=13, value=row.get("Texto do Item", ""))  # Coluna I
            ws.cell(row=i, column=16, value=row.get("Observação (Texto da Situação)", ""))  # Coluna J
        except KeyError as e:
            raise KeyError(f"Chave ausente nos dados: {str(e)}")

    # Salva o arquivo preenchido em memória
    output = io.BytesIO()
    try:
        wb.save(output)
        output.seek(0)
    except Exception as e:
        raise Exception(f"Erro ao salvar o arquivo Excel: {str(e)}")

    return output

@app.route('/finalize', methods=['GET'])
def finalize():
    try:
        output = generate_excel()
        # Usa o valor de LT como parte do nome do arquivo
        lt_name = default_values.get("LT", "inspecao").strip()  # Remove espaços extras
        if not lt_name:  # Caso LT esteja vazio
            lt_name = "inspecao"
        file_name = f"{lt_name}.xlsx"
        return send_file(output, as_attachment=True, download_name=file_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except FileNotFoundError as e:
        return jsonify({"status": "error", "message": str(e)}), 400
    except KeyError as e:
        return jsonify({"status": "error", "message": f"Erro nos dados: {str(e)}"}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": f"Erro ao gerar o arquivo Excel: {str(e)}"}), 500

@app.route('/reset', methods=['POST'])
def reset():
    global data_rows
    data_rows = []  # Limpa os dados em memória

    # Limpa o arquivo data.json
    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data_rows, f, ensure_ascii=False, indent=4)

    return jsonify({"status": "success", "message": "Dados reiniciados com sucesso!"})

def save_data_to_file():
    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data_rows, f, ensure_ascii=False, indent=4)

def load_data_from_file():
    global data_rows
    if os.path.exists("data.json"):
        with open("data.json", "r", encoding="utf-8") as f:
            data_rows = json.load(f)

if __name__ == '__main__':
    debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=5005, debug=debug_mode)

[]
